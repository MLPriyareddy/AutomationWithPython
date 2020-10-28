import openpyxl


def getrowcount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_row


def getcolumncount(path, sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.max_column


def readdata(path, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rownum, column=columnno).value


def writedata(path, sheetname, rownum, columno, data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rownum, column=columno).value = data
    workbook.save(path)

